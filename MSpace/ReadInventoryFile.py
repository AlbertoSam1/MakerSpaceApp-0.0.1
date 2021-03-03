import openpyxl
from FileManager import open_single_file_manager

# Read and check a CSV file
class CheckCSV:
    def __init__(self):
        self.file = open_single_file_manager()

    def later(self):
        xlsx_file = Path(Path.home(), 'Documents', 'ExcelTest.xlsx')
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        keys = []

        for column in sheet.iter_cols(1, sheet.max_column):
            keys.append(column[0].value)

        values = [list(row) for row
                  in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True)]

        file_dict = dict(zip(keys, values))
        print(file_dict)
